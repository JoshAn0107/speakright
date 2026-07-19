from fastapi import APIRouter, Depends, HTTPException, status, UploadFile, File, Form
from sqlalchemy.orm import Session
from sqlalchemy import func, and_
from typing import List, Optional
from datetime import datetime
from pathlib import Path
import os
import re
import shutil

from app.db.session import get_db
from app.api.deps import get_current_teacher, get_current_student, get_current_user
from app.models.user import User
from app.models.assignment import (
    WordDatabase, WordDatabaseWord, Assignment, AssignmentWord,
    AssignmentStudent, AssignmentSubmission
)
from app.models.recording import Recording, RecordingStatus
from app.models.wordlist_upload import WordlistUpload
from app.schemas.assignment import (
    WordDatabaseResponse, WordDatabaseWordResponse,
    WordDatabaseCreate, WordDatabaseWordsBulkCreate,
    AssignmentCreate, AssignmentUpdate, AssignmentResponse,
    StudentAssignmentResponse, AssignmentProgressResponse,
    AssignmentSubmissionCreate
)

router = APIRouter()


# ===== Word Database Endpoints =====

@router.get("/databases", response_model=List[WordDatabaseResponse])
def get_word_databases(
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user)
):
    """Get all available word databases (excluding those in the recycle bin)"""
    databases = db.query(WordDatabase).filter(WordDatabase.deleted_at.is_(None)).all()
    return databases


@router.get("/databases/{database_id}/words", response_model=List[WordDatabaseWordResponse])
def get_database_words(
    database_id: int,
    skip: int = 0,
    limit: int = 1000,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user)
):
    """Get words from a specific database"""
    database = db.query(WordDatabase).filter(WordDatabase.id == database_id).first()
    if not database:
        raise HTTPException(status_code=404, detail="未找到词库")

    words = db.query(WordDatabaseWord).filter(
        WordDatabaseWord.database_id == database_id
    ).offset(skip).limit(limit).all()

    return words


def _get_owned_database(database_id: int, db: Session, current_user: User) -> WordDatabase:
    database = db.query(WordDatabase).filter(WordDatabase.id == database_id).first()
    if not database:
        raise HTTPException(status_code=404, detail="未找到词库")
    if database.created_by != current_user.id:
        raise HTTPException(status_code=403, detail="只能修改自己创建的词库")
    return database


@router.post("/databases", response_model=WordDatabaseResponse)
def create_word_database(
    database_data: WordDatabaseCreate,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_teacher)
):
    """Teacher creates a custom word database (e.g. 上海新教材六年级上)"""
    name = database_data.name.strip()
    if not name:
        raise HTTPException(status_code=400, detail="词库名称不能为空")

    existing = db.query(WordDatabase).filter(WordDatabase.name == name).first()
    if existing:
        raise HTTPException(status_code=400, detail="已存在同名词库")

    database = WordDatabase(
        name=name,
        description=database_data.description,
        word_count=0,
        created_by=current_user.id
    )
    db.add(database)
    db.commit()
    db.refresh(database)
    return database


@router.post("/databases/{database_id}/words", response_model=dict)
def add_words_to_database(
    database_id: int,
    payload: WordDatabaseWordsBulkCreate,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_teacher)
):
    """Add words (bulk) to a teacher's own word database; duplicates are skipped"""
    database = _get_owned_database(database_id, db, current_user)

    existing_words = {
        w.word_text for w in db.query(WordDatabaseWord).filter(
            WordDatabaseWord.database_id == database_id
        ).all()
    }

    added = 0
    skipped = 0
    for item in payload.words:
        word_text = item.word_text.lower().strip()
        if not word_text or word_text in existing_words:
            skipped += 1
            continue
        db.add(WordDatabaseWord(
            database_id=database_id,
            word_text=word_text,
            definition=item.definition,
            example_sentence=item.example_sentence,
            difficulty_level=item.difficulty_level,
            unit=(item.unit or "").strip() or None
        ))
        existing_words.add(word_text)
        added += 1

    database.word_count = len(existing_words)
    db.commit()

    return {
        "message": f"已添加{added}个单词" + (f"，跳过{skipped}个（重复或为空）" if skipped else ""),
        "added": added,
        "skipped": skipped,
        "word_count": database.word_count
    }


@router.delete("/databases/{database_id}/words/{word_id}", response_model=dict)
def delete_database_word(
    database_id: int,
    word_id: int,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_teacher)
):
    """Remove a word from a teacher's own word database"""
    database = _get_owned_database(database_id, db, current_user)

    word = db.query(WordDatabaseWord).filter(
        WordDatabaseWord.id == word_id,
        WordDatabaseWord.database_id == database_id
    ).first()
    if not word:
        raise HTTPException(status_code=404, detail="未找到单词")

    db.delete(word)
    database.word_count = max(0, (database.word_count or 1) - 1)
    db.commit()

    return {"message": "单词已删除", "word_count": database.word_count}


@router.get("/databases/trash", response_model=List[dict])
def get_trashed_databases(
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_teacher)
):
    """List the teacher's word databases currently in the recycle bin"""
    databases = db.query(WordDatabase).filter(
        WordDatabase.created_by == current_user.id,
        WordDatabase.deleted_at.isnot(None)
    ).order_by(WordDatabase.deleted_at.desc()).all()

    return [
        {
            "id": d.id,
            "name": d.name,
            "description": d.description,
            "word_count": d.word_count,
            "deleted_at": d.deleted_at
        }
        for d in databases
    ]


@router.delete("/databases/trash", response_model=dict)
def empty_trash(
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_teacher)
):
    """Permanently delete all of the teacher's databases in the recycle bin"""
    trashed = db.query(WordDatabase).filter(
        WordDatabase.created_by == current_user.id,
        WordDatabase.deleted_at.isnot(None)
    ).all()

    for database in trashed:
        db.query(WordDatabaseWord).filter(
            WordDatabaseWord.database_id == database.id
        ).delete()
        db.delete(database)
    db.commit()

    return {"message": f"回收站已清空，彻底删除了{len(trashed)}个词库", "purged": len(trashed)}


@router.post("/databases/{database_id}/restore", response_model=dict)
def restore_database(
    database_id: int,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_teacher)
):
    """Restore a word database from the recycle bin"""
    database = _get_owned_database(database_id, db, current_user)
    if database.deleted_at is None:
        raise HTTPException(status_code=400, detail="该词库不在回收站中")

    database.deleted_at = None
    db.commit()

    return {"message": "词库已恢复", "database_id": database_id, "name": database.name}


@router.delete("/databases/{database_id}/purge", response_model=dict)
def purge_database(
    database_id: int,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_teacher)
):
    """Permanently delete a database from the recycle bin (cannot be undone)"""
    database = _get_owned_database(database_id, db, current_user)
    if database.deleted_at is None:
        raise HTTPException(status_code=400, detail="请先将词库放入回收站再彻底删除")

    db.query(WordDatabaseWord).filter(
        WordDatabaseWord.database_id == database_id
    ).delete()
    db.delete(database)
    db.commit()

    return {"message": "词库已彻底删除", "database_id": database_id}


@router.delete("/databases/{database_id}", response_model=dict)
def delete_word_database(
    database_id: int,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_teacher)
):
    """Move a teacher's own word database to the recycle bin (soft delete)"""
    database = _get_owned_database(database_id, db, current_user)

    database.deleted_at = datetime.utcnow()
    db.commit()

    return {"message": "词库已放入回收站，可在回收站中恢复", "database_id": database_id}


@router.get("/databases/{database_id}/export")
def export_database_excel(
    database_id: int,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user)
):
    """Download a word database as an Excel (.xlsx) file"""
    from io import BytesIO
    from urllib.parse import quote
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill
    from openpyxl.utils import get_column_letter
    from fastapi.responses import StreamingResponse

    database = db.query(WordDatabase).filter(WordDatabase.id == database_id).first()
    if not database:
        raise HTTPException(status_code=404, detail="未找到词库")

    words = db.query(WordDatabaseWord).filter(
        WordDatabaseWord.database_id == database_id
    ).order_by(WordDatabaseWord.unit, WordDatabaseWord.id).all()

    wb = Workbook()
    ws = wb.active
    ws.title = database.name[:31]  # Excel sheet name limit

    headers = ["组别", "单词", "释义", "例句"]
    ws.append(headers)
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    for cell in ws[1]:
        cell.font = header_font
        cell.fill = header_fill

    for w in words:
        ws.append([w.unit or "", w.word_text, w.definition or "", w.example_sentence or ""])

    for col, width in enumerate([12, 24, 50, 50], start=1):
        ws.column_dimensions[get_column_letter(col)].width = width
    ws.freeze_panes = "A2"

    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)

    filename = f"{database.name}.xlsx"
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename*=UTF-8''{quote(filename)}"}
    )


# ===== Wordlist Upload (等待适配) Endpoints =====

WORDLIST_UPLOAD_DIR = Path("wordlist_uploads")
WORDLIST_UPLOAD_DIR.mkdir(exist_ok=True)
WORDLIST_ALLOWED_EXTENSIONS = {
    ".txt", ".csv", ".md", ".xlsx", ".xls", ".docx", ".doc",
    ".pdf", ".png", ".jpg", ".jpeg", ".webp"
}
WORDLIST_MAX_SIZE = 20 * 1024 * 1024  # 20MB


@router.post("/wordlist-uploads", response_model=dict)
def upload_wordlist_file(
    file: UploadFile = File(...),
    target_name: Optional[str] = Form(None),
    note: Optional[str] = Form(None),
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_teacher)
):
    """Teacher uploads a wordlist file (txt/Excel/PDF/photo) to be adapted
    into a word database by the operator."""

    extension = os.path.splitext(file.filename or "")[1].lower()
    if extension not in WORDLIST_ALLOWED_EXTENSIONS:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail=f"不支持的文件类型 {extension}，支持：txt、csv、Excel、Word、PDF、图片"
        )

    upload = WordlistUpload(
        teacher_id=current_user.id,
        original_filename=file.filename,
        stored_path="",
        target_name=(target_name or "").strip() or None,
        note=(note or "").strip() or None,
        status="pending"
    )
    db.add(upload)
    db.flush()  # get id for the filename

    safe_name = re.sub(r"[^\w.\-\u4e00-\u9fff]", "_", file.filename or "file")
    stored_path = WORDLIST_UPLOAD_DIR / f"{upload.id}_{safe_name}"

    size = 0
    try:
        with open(stored_path, "wb") as buffer:
            while chunk := file.file.read(1024 * 1024):
                size += len(chunk)
                if size > WORDLIST_MAX_SIZE:
                    raise HTTPException(
                        status_code=status.HTTP_400_BAD_REQUEST,
                        detail="文件不能超过20MB"
                    )
                buffer.write(chunk)
    except HTTPException:
        db.rollback()
        if stored_path.exists():
            stored_path.unlink()
        raise
    except Exception as e:
        db.rollback()
        if stored_path.exists():
            stored_path.unlink()
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail=f"无法保存文件：{str(e)}"
        )

    upload.stored_path = str(stored_path)
    db.commit()

    return {
        "message": "文件已上传，等待适配。适配完成后词库会自动出现在词库列表中。",
        "upload_id": upload.id,
        "original_filename": upload.original_filename,
        "status": upload.status
    }


@router.get("/wordlist-uploads", response_model=List[dict])
def get_my_wordlist_uploads(
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_teacher)
):
    """List the teacher's uploaded wordlist files and their adaptation status"""

    uploads = db.query(WordlistUpload).filter(
        WordlistUpload.teacher_id == current_user.id
    ).order_by(WordlistUpload.created_at.desc()).all()

    return [
        {
            "id": u.id,
            "original_filename": u.original_filename,
            "target_name": u.target_name,
            "note": u.note,
            "status": u.status,
            "result_message": u.result_message,
            "created_at": u.created_at,
            "processed_at": u.processed_at
        }
        for u in uploads
    ]


@router.delete("/wordlist-uploads/{upload_id}", response_model=dict)
def delete_wordlist_upload(
    upload_id: int,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_teacher)
):
    """Teacher deletes their own upload (only while still pending)"""

    upload = db.query(WordlistUpload).filter(
        WordlistUpload.id == upload_id,
        WordlistUpload.teacher_id == current_user.id
    ).first()

    if not upload:
        raise HTTPException(status_code=404, detail="未找到上传记录")

    if upload.status != "pending":
        raise HTTPException(status_code=400, detail="该文件已适配，无法删除记录")

    if upload.stored_path and os.path.exists(upload.stored_path):
        os.remove(upload.stored_path)
    db.delete(upload)
    db.commit()

    return {"message": "上传已删除", "upload_id": upload_id}


# ===== Teacher Assignment Endpoints =====

@router.post("/teacher/assignments", response_model=AssignmentResponse)
def create_assignment(
    assignment_data: AssignmentCreate,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_teacher)
):
    """Teacher creates a new assignment"""

    # Validate word count (20-40 words)
    word_count = len(assignment_data.words)
    if word_count < 20 or word_count > 40:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail=f"作业必须包含20-40个单词。当前为{word_count}个。"
        )

    # Validate student IDs
    if not assignment_data.student_ids:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail="至少需要分配一名学生"
        )

    students = db.query(User).filter(
        User.id.in_(assignment_data.student_ids),
        User.role == "student"
    ).all()

    if len(students) != len(assignment_data.student_ids):
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail="一个或多个学生ID无效"
        )

    # Teachers may only assign to students enrolled in their own classes
    from app.models.classes import Class, ClassEnrollment
    own_student_ids = {
        r[0] for r in db.query(ClassEnrollment.student_id).join(
            Class, ClassEnrollment.class_id == Class.id
        ).filter(Class.teacher_id == current_user.id).all()
    }
    outside = [sid for sid in assignment_data.student_ids if sid not in own_student_ids]
    if outside:
        raise HTTPException(
            status_code=status.HTTP_403_FORBIDDEN,
            detail="只能给自己班级的学生布置作业，请先让学生用班级码加入你的班级"
        )

    try:
        # Create assignment
        assignment = Assignment(
            teacher_id=current_user.id,
            title=assignment_data.title,
            description=assignment_data.description,
            mode=assignment_data.mode if assignment_data.mode in ("practice", "continuous") else "practice",
            word_database_id=assignment_data.word_database_id,
            due_date=assignment_data.due_date
        )
        db.add(assignment)
        db.flush()  # Get assignment ID

        # Add words to assignment
        for index, word_text in enumerate(assignment_data.words):
            assignment_word = AssignmentWord(
                assignment_id=assignment.id,
                word_text=word_text.lower().strip(),
                order_index=index
            )
            db.add(assignment_word)

        # Assign to students
        for student_id in assignment_data.student_ids:
            assignment_student = AssignmentStudent(
                assignment_id=assignment.id,
                student_id=student_id
            )
            db.add(assignment_student)

        db.commit()
        db.refresh(assignment)
    except Exception as e:
        db.rollback()
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail=f"创建作业时数据库错误：{str(e)}"
        )

    # Build response
    return build_assignment_response(assignment, db)


@router.get("/teacher/assignments", response_model=List[AssignmentResponse])
def get_teacher_assignments(
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_teacher)
):
    """Get all assignments created by the teacher"""
    assignments = db.query(Assignment).filter(
        Assignment.teacher_id == current_user.id
    ).order_by(Assignment.created_at.desc()).all()

    return [build_assignment_response(assignment, db) for assignment in assignments]


@router.get("/teacher/assignments/{assignment_id}", response_model=AssignmentResponse)
def get_teacher_assignment(
    assignment_id: int,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_teacher)
):
    """Get a specific assignment created by the teacher"""
    assignment = db.query(Assignment).filter(
        Assignment.id == assignment_id,
        Assignment.teacher_id == current_user.id
    ).first()

    if not assignment:
        raise HTTPException(status_code=404, detail="未找到作业")

    return build_assignment_response(assignment, db)


@router.put("/teacher/assignments/{assignment_id}", response_model=AssignmentResponse)
def update_assignment(
    assignment_id: int,
    assignment_data: AssignmentUpdate,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_teacher)
):
    """Update an assignment"""
    assignment = db.query(Assignment).filter(
        Assignment.id == assignment_id,
        Assignment.teacher_id == current_user.id
    ).first()

    if not assignment:
        raise HTTPException(status_code=404, detail="未找到作业")

    # Update fields
    if assignment_data.title is not None:
        assignment.title = assignment_data.title
    if assignment_data.description is not None:
        assignment.description = assignment_data.description
    if assignment_data.due_date is not None:
        assignment.due_date = assignment_data.due_date

    db.commit()
    db.refresh(assignment)

    return build_assignment_response(assignment, db)


@router.delete("/teacher/assignments/{assignment_id}")
def delete_assignment(
    assignment_id: int,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_teacher)
):
    """Delete an assignment"""
    assignment = db.query(Assignment).filter(
        Assignment.id == assignment_id,
        Assignment.teacher_id == current_user.id
    ).first()

    if not assignment:
        raise HTTPException(status_code=404, detail="未找到作业")

    db.delete(assignment)
    db.commit()

    return {"message": "作业删除成功"}


@router.get("/teacher/assignments/{assignment_id}/progress", response_model=List[dict])
def get_assignment_progress(
    assignment_id: int,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_teacher)
):
    """Get detailed progress for all students on an assignment"""
    assignment = db.query(Assignment).filter(
        Assignment.id == assignment_id,
        Assignment.teacher_id == current_user.id
    ).first()

    if not assignment:
        raise HTTPException(status_code=404, detail="未找到作业")

    total_words = len(assignment.words)
    student_progress = []

    # one aggregate query for all students: submission count + average score
    agg = {
        r[0]: (r[1], r[2])
        for r in db.query(
            AssignmentSubmission.student_id,
            func.count(AssignmentSubmission.id),
            func.avg(func.nullif(func.json_extract(Recording.automated_scores, "$.pronunciation_score"), 0))
        ).outerjoin(
            Recording, AssignmentSubmission.recording_id == Recording.id
        ).filter(
            AssignmentSubmission.assignment_id == assignment_id
        ).group_by(AssignmentSubmission.student_id).all()
    }

    for assignment_student in assignment.students:
        student = assignment_student.student
        completed_count, avg_raw = agg.get(student.id, (0, None))
        average_score = round(float(avg_raw), 1) if avg_raw is not None else None

        completion_percentage = (completed_count / total_words * 100) if total_words > 0 else 0

        student_progress.append({
            "student_id": student.id,
            "student_name": student.username,
            "total_words": total_words,
            "completed_words": completed_count,
            "completion_percentage": round(completion_percentage, 1),
            "average_score": average_score,
            "assigned_at": assignment_student.assigned_at,
            "completed_at": assignment_student.completed_at
        })

    return student_progress


@router.post("/teacher/assignments/{assignment_id}/students/{student_id}/word-feedback")
def submit_word_feedback(
    assignment_id: int,
    student_id: int,
    word_text: str,
    feedback_text: str = "",
    grade: str = "",
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_teacher)
):
    """Per-word teacher feedback — works for continuous takes where all
    words share one recording (recording-level feedback would overwrite)."""
    assignment = db.query(Assignment).filter(
        Assignment.id == assignment_id,
        Assignment.teacher_id == current_user.id
    ).first()
    if not assignment:
        raise HTTPException(status_code=404, detail="未找到作业")

    sub = db.query(AssignmentSubmission).filter(
        AssignmentSubmission.assignment_id == assignment_id,
        AssignmentSubmission.student_id == student_id,
        AssignmentSubmission.word_text == word_text
    ).order_by(AssignmentSubmission.id.desc()).first()
    if not sub:
        raise HTTPException(status_code=404, detail="该学生尚未提交这个单词")

    sub.teacher_feedback = feedback_text.strip() or None
    sub.teacher_grade = grade.strip() or None
    db.commit()
    return {"message": "点评已保存", "word_text": word_text}


@router.get("/teacher/assignments/{assignment_id}/students/{student_id}/progress")
def get_student_progress_for_teacher(
    assignment_id: int,
    student_id: int,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_teacher)
):
    """Get detailed word-by-word progress for a specific student on an assignment (teacher view)"""
    # Verify teacher owns this assignment
    assignment = db.query(Assignment).filter(
        Assignment.id == assignment_id,
        Assignment.teacher_id == current_user.id
    ).first()

    if not assignment:
        raise HTTPException(status_code=404, detail="未找到作业")

    # Verify student is assigned to this assignment
    assignment_student = db.query(AssignmentStudent).filter(
        AssignmentStudent.assignment_id == assignment_id,
        AssignmentStudent.student_id == student_id
    ).first()

    if not assignment_student:
        raise HTTPException(status_code=404, detail="学生未被分配到此作业")

    # Get all words in assignment
    assignment_words = db.query(AssignmentWord).filter(
        AssignmentWord.assignment_id == assignment_id
    ).order_by(AssignmentWord.order_index).all()

    # Get submissions for this student
    submissions = db.query(AssignmentSubmission).filter(
        AssignmentSubmission.assignment_id == assignment_id,
        AssignmentSubmission.student_id == student_id
    ).all()

    submission_dict = {sub.word_text: sub for sub in submissions}

    mode = assignment.mode or "practice"

    # continuous mode: every word shares one take — use its per-word breakdown
    continuous_summary = None
    per_word_map = {}
    if mode == "continuous":
        latest_sub = max((s for s in submissions if s.recording_id), key=lambda x: x.id, default=None)
        if latest_sub and latest_sub.recording:
            rec = latest_sub.recording
            scores = rec.automated_scores or {}
            per_word_map = {w.get("word"): w for w in scores.get("per_word", [])}
            continuous_summary = {
                "recording_id": rec.id,
                "audio_file_path": rec.audio_file_path,
                "pronunciation_score": scores.get("pronunciation_score"),
                "grade": rec.teacher_grade,
                "feedback": rec.teacher_feedback,
                "words_read": scores.get("words_read"),
                "words_total": scores.get("words_total"),
                "completeness_score": scores.get("completeness_score"),
                "fluency_score": scores.get("fluency_score"),
                "recognized_text": scores.get("recognized_text"),
                "scoring": scores == {} or (rec.status == RecordingStatus.PENDING and not scores),
            }

    word_list = []
    for word in assignment_words:
        submission = submission_dict.get(word.word_text)
        word_info = {
            "word_text": word.word_text,
            "order_index": word.order_index,
            "submitted": submission is not None,
            "recording_id": submission.recording_id if submission else None,
            "submitted_at": submission.submitted_at if submission else None,
            "teacher_feedback": submission.teacher_feedback if submission else None,
            "teacher_grade": submission.teacher_grade if submission else None
        }

        if mode == "continuous":
            pw = per_word_map.get(word.word_text)
            if pw:
                word_info["score"] = pw.get("score")
                word_info["error"] = pw.get("error")
                word_info["offset_ms"] = pw.get("offset_ms")
                word_info["end_ms"] = pw.get("end_ms")
        elif submission and submission.recording_id:
            recording = db.query(Recording).filter(Recording.id == submission.recording_id).first()
            if recording and recording.automated_scores:
                word_info["score"] = recording.automated_scores.get("pronunciation_score", 0)

        word_list.append(word_info)

    total_words = len(assignment_words)
    completed_words = len([w for w in word_list if w["submitted"]])

    return {
        "assignment_id": assignment_id,
        "assignment_title": assignment.title,
        "mode": mode,
        "continuous_summary": continuous_summary,
        "student_id": student_id,
        "total_words": total_words,
        "completed_words": completed_words,
        "completion_percentage": round((completed_words / total_words * 100), 1) if total_words > 0 else 0,
        "completed_word_texts": [w["word_text"] for w in word_list if w["submitted"]],
        "words": word_list
    }


# ===== Student Assignment Endpoints =====

@router.get("/student/assignments", response_model=List[StudentAssignmentResponse])
def get_student_assignments(
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_student)
):
    """Get all assignments for the current student"""
    assignment_students = db.query(AssignmentStudent).filter(
        AssignmentStudent.student_id == current_user.id
    ).all()

    assignments = []
    for assignment_student in assignment_students:
        assignment = assignment_student.assignment
        assignments.append(build_student_assignment_response(assignment, current_user.id, db))

    return assignments


@router.get("/student/assignments/{assignment_id}", response_model=StudentAssignmentResponse)
def get_student_assignment(
    assignment_id: int,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_student)
):
    """Get a specific assignment for the student"""
    assignment_student = db.query(AssignmentStudent).filter(
        AssignmentStudent.assignment_id == assignment_id,
        AssignmentStudent.student_id == current_user.id
    ).first()

    if not assignment_student:
        raise HTTPException(status_code=404, detail="未找到作业")

    assignment = assignment_student.assignment
    return build_student_assignment_response(assignment, current_user.id, db)


@router.post("/student/assignments/{assignment_id}/submit-continuous")
def submit_assignment_continuous(
    assignment_id: int,
    audio_file: UploadFile = File(...),
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_student)
):
    """Test mode: accept the take immediately, score it in the background."""
    from app.services.shadow_service import submit_shadow
    from app.models.recording import Recording, RecordingStatus
    from app.core.config import settings as app_settings

    assignment_student = db.query(AssignmentStudent).filter(
        AssignmentStudent.assignment_id == assignment_id,
        AssignmentStudent.student_id == current_user.id
    ).first()
    if not assignment_student:
        raise HTTPException(status_code=404, detail="未找到作业")

    assignment = assignment_student.assignment
    reference_words = [w.word_text for w in sorted(assignment.words, key=lambda x: x.order_index)]
    if not reference_words:
        raise HTTPException(status_code=400, detail="作业没有单词")

    upload_dir = Path(app_settings.UPLOAD_DIR) / str(current_user.id)
    upload_dir.mkdir(parents=True, exist_ok=True)
    ext = os.path.splitext(audio_file.filename or "")[1] or ".wav"
    file_path = upload_dir / f"continuous_{assignment_id}_{datetime.now().strftime('%Y%m%d_%H%M%S')}{ext}"
    with open(file_path, "wb") as buffer:
        import shutil as _shutil
        _shutil.copyfileobj(audio_file.file, buffer)

    # placeholder recording: student doesn't wait for scoring
    recording = Recording(
        student_id=current_user.id,
        word_text=f"[连读] {assignment.title}"[:100],
        audio_file_path=str(file_path),
        automated_scores=None,
        teacher_feedback="评分中…",
        teacher_grade=None,
        status=RecordingStatus.PENDING,
    )
    db.add(recording)
    db.flush()

    # submissions recorded now so progress reflects completion immediately
    db.query(AssignmentSubmission).filter(
        AssignmentSubmission.assignment_id == assignment_id,
        AssignmentSubmission.student_id == current_user.id
    ).delete()
    for word in reference_words:
        db.add(AssignmentSubmission(
            assignment_id=assignment_id,
            student_id=current_user.id,
            word_text=word,
            recording_id=recording.id
        ))
    assignment_student.completed_at = datetime.utcnow()
    db.commit()
    recording_id = recording.id

    def score_in_background():
        from app.db.session import SessionLocal
        from app.services.pronunciation_service import pronunciation_service
        from app.services.feedback_service import FeedbackService
        session = SessionLocal()
        try:
            result = pronunciation_service.assess_continuous_reading(str(file_path), reference_words)
            rec = session.query(Recording).filter(Recording.id == recording_id).first()
            if not rec:
                return
            if result.get("error"):
                rec.automated_scores = {"error": result["error"], "pronunciation_score": 0, "per_word": []}
                rec.teacher_feedback = "自动评分失败，可以重新测试，或等老师人工评分。"
                rec.status = RecordingStatus.PENDING
            else:
                overall = result.get("pronunciation_score", 0)
                grade = FeedbackService._calculate_grade(overall)
                missed = [w["word"] for w in result.get("per_word", []) if w.get("error") == "漏读"]
                weak = [w["word"] for w in result.get("per_word", []) if w.get("error") != "漏读" and w.get("score", 100) < 60]
                parts = [f"连读测试完成：{result.get('words_read', 0)}/{result.get('words_total', 0)} 个单词，总分 {overall:.0f}。"]
                if missed:
                    parts.append(f"漏读：{'、'.join(missed[:10])}{'…' if len(missed) > 10 else ''}。")
                if weak:
                    parts.append(f"发音需加强：{'、'.join(weak[:10])}{'…' if len(weak) > 10 else ''}。")
                if not missed and not weak:
                    parts.append("全部单词都读到位了，很棒！")
                rec.automated_scores = result
                rec.teacher_feedback = " ".join(parts)
                rec.teacher_grade = grade
                rec.status = RecordingStatus.REVIEWED
                rec.reviewed_at = datetime.utcnow()
            session.commit()
            try:
                submit_shadow(recording_id, " ".join(reference_words), str(file_path), result if not result.get("error") else {})
            except Exception as e:
                print(f"Shadow submit failed (non-fatal): {e}")
        except Exception as e:
            import traceback
            traceback.print_exc()
        finally:
            session.close()

    import threading
    threading.Thread(target=score_in_background, daemon=True, name=f"continuous-score-{recording_id}").start()

    return {
        "message": "已提交，正在后台评分",
        "recording_id": recording_id,
        "status": "scoring"
    }


@router.get("/student/assignments/{assignment_id}/continuous-result")
def get_continuous_result(
    assignment_id: int,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_student)
):
    """Poll the latest continuous-test result for this assignment."""
    from app.models.recording import Recording, RecordingStatus

    sub = db.query(AssignmentSubmission).filter(
        AssignmentSubmission.assignment_id == assignment_id,
        AssignmentSubmission.student_id == current_user.id,
        AssignmentSubmission.recording_id.isnot(None)
    ).order_by(AssignmentSubmission.id.desc()).first()
    if not sub or not sub.recording:
        return {"status": "none"}

    rec = sub.recording
    scores = rec.automated_scores or {}
    if rec.status == RecordingStatus.PENDING and not scores:
        return {"status": "scoring", "message": "评分中，请稍后刷新"}
    if scores.get("error"):
        return {"status": "failed", "message": "这次没评好，可以重新测试，或等老师人工评分"}

    fb_map = {
        x.word_text: (x.teacher_feedback, x.teacher_grade)
        for x in db.query(AssignmentSubmission).filter(
            AssignmentSubmission.assignment_id == assignment_id,
            AssignmentSubmission.student_id == current_user.id
        ).all()
        if x.teacher_feedback or x.teacher_grade
    }
    per_word = []
    for w in scores.get("per_word", []):
        w = dict(w)
        fb = fb_map.get(w.get("word"))
        if fb:
            w["teacher_feedback"], w["teacher_grade"] = fb
        per_word.append(w)

    return {
        "status": "done",
        "recording_id": rec.id,
        "pronunciation_score": scores.get("pronunciation_score"),
        "grade": rec.teacher_grade,
        "feedback": rec.teacher_feedback,
        "per_word": per_word,
        "words_read": scores.get("words_read"),
        "words_total": scores.get("words_total"),
        "completeness_score": scores.get("completeness_score"),
        "fluency_score": scores.get("fluency_score"),
    }


@router.post("/student/assignments/{assignment_id}/submit")
def submit_assignment_word(
    assignment_id: int,
    word_text: str,
    recording_id: int,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_student)
):
    """Submit a word recording for an assignment"""
    # Verify student has this assignment
    assignment_student = db.query(AssignmentStudent).filter(
        AssignmentStudent.assignment_id == assignment_id,
        AssignmentStudent.student_id == current_user.id
    ).first()

    if not assignment_student:
        raise HTTPException(status_code=404, detail="未找到作业")

    # Verify word is in assignment
    assignment_word = db.query(AssignmentWord).filter(
        AssignmentWord.assignment_id == assignment_id,
        AssignmentWord.word_text == word_text.lower().strip()
    ).first()

    if not assignment_word:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail="单词不在作业中"
        )

    # Verify recording belongs to student
    recording = db.query(Recording).filter(
        Recording.id == recording_id,
        Recording.student_id == current_user.id
    ).first()

    if not recording:
        raise HTTPException(status_code=404, detail="未找到录音")

    # Check if submission already exists (allow re-recording)
    existing_submission = db.query(AssignmentSubmission).filter(
        AssignmentSubmission.assignment_id == assignment_id,
        AssignmentSubmission.student_id == current_user.id,
        AssignmentSubmission.word_text == word_text.lower().strip()
    ).first()

    if existing_submission:
        # Update existing submission
        existing_submission.recording_id = recording_id
        existing_submission.submitted_at = datetime.utcnow()
    else:
        # Create new submission
        submission = AssignmentSubmission(
            assignment_id=assignment_id,
            student_id=current_user.id,
            word_text=word_text.lower().strip(),
            recording_id=recording_id
        )
        db.add(submission)

    # Check if assignment is complete
    total_words = db.query(AssignmentWord).filter(
        AssignmentWord.assignment_id == assignment_id
    ).count()

    completed_words = db.query(AssignmentSubmission).filter(
        AssignmentSubmission.assignment_id == assignment_id,
        AssignmentSubmission.student_id == current_user.id
    ).count()

    if completed_words >= total_words and not assignment_student.completed_at:
        assignment_student.completed_at = datetime.utcnow()

    db.commit()

    return {
        "message": "单词提交成功",
        "completed_words": completed_words,
        "total_words": total_words,
        "completion_percentage": round((completed_words / total_words * 100), 1) if total_words > 0 else 0
    }


@router.get("/student/assignments/{assignment_id}/progress")
def get_student_assignment_progress(
    assignment_id: int,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_student)
):
    """Get student's progress on a specific assignment"""
    assignment_student = db.query(AssignmentStudent).filter(
        AssignmentStudent.assignment_id == assignment_id,
        AssignmentStudent.student_id == current_user.id
    ).first()

    if not assignment_student:
        raise HTTPException(status_code=404, detail="未找到作业")

    assignment = assignment_student.assignment

    # Get all words in assignment
    assignment_words = db.query(AssignmentWord).filter(
        AssignmentWord.assignment_id == assignment_id
    ).order_by(AssignmentWord.order_index).all()

    # Get submissions
    submissions = db.query(AssignmentSubmission).filter(
        AssignmentSubmission.assignment_id == assignment_id,
        AssignmentSubmission.student_id == current_user.id
    ).all()

    submission_dict = {sub.word_text: sub for sub in submissions}

    mode = assignment.mode or "practice"

    # continuous mode: every word shares one take — use its per-word breakdown
    continuous_summary = None
    per_word_map = {}
    if mode == "continuous":
        latest_sub = max((s for s in submissions if s.recording_id), key=lambda x: x.id, default=None)
        if latest_sub and latest_sub.recording:
            rec = latest_sub.recording
            scores = rec.automated_scores or {}
            per_word_map = {w.get("word"): w for w in scores.get("per_word", [])}
            continuous_summary = {
                "recording_id": rec.id,
                "audio_file_path": rec.audio_file_path,
                "pronunciation_score": scores.get("pronunciation_score"),
                "grade": rec.teacher_grade,
                "feedback": rec.teacher_feedback,
                "words_read": scores.get("words_read"),
                "words_total": scores.get("words_total"),
                "completeness_score": scores.get("completeness_score"),
                "fluency_score": scores.get("fluency_score"),
                "recognized_text": scores.get("recognized_text"),
                "scoring": scores == {} or (rec.status == RecordingStatus.PENDING and not scores),
            }

    word_list = []
    for word in assignment_words:
        submission = submission_dict.get(word.word_text)
        word_info = {
            "word_text": word.word_text,
            "order_index": word.order_index,
            "submitted": submission is not None,
            "recording_id": submission.recording_id if submission else None,
            "submitted_at": submission.submitted_at if submission else None,
            "teacher_feedback": submission.teacher_feedback if submission else None,
            "teacher_grade": submission.teacher_grade if submission else None
        }

        if mode == "continuous":
            pw = per_word_map.get(word.word_text)
            if pw:
                word_info["score"] = pw.get("score")
                word_info["error"] = pw.get("error")
                word_info["offset_ms"] = pw.get("offset_ms")
                word_info["end_ms"] = pw.get("end_ms")
        elif submission and submission.recording_id:
            recording = db.query(Recording).filter(Recording.id == submission.recording_id).first()
            if recording and recording.automated_scores:
                word_info["score"] = recording.automated_scores.get("pronunciation_score", 0)

        word_list.append(word_info)

    total_words = len(assignment_words)
    completed_words = len([w for w in word_list if w["submitted"]])

    return {
        "assignment_id": assignment_id,
        "assignment_title": assignment.title,
        "total_words": total_words,
        "completed_words": completed_words,
        "completion_percentage": round((completed_words / total_words * 100), 1) if total_words > 0 else 0,
        "words": word_list
    }


# ===== Helper Functions =====

def build_assignment_response(assignment: Assignment, db: Session) -> dict:
    """Build assignment response with additional data"""
    word_count = len(assignment.words)
    student_count = len(assignment.students)

    # Get word database name if exists
    word_database_name = None
    if assignment.word_database_id:
        database = db.query(WordDatabase).filter(WordDatabase.id == assignment.word_database_id).first()
        if database:
            word_database_name = database.name

    return {
        "id": assignment.id,
        "teacher_id": assignment.teacher_id,
        "title": assignment.title,
        "description": assignment.description,
        "mode": assignment.mode or "practice",
        "word_database_id": assignment.word_database_id,
        "word_database_name": word_database_name,
        "due_date": assignment.due_date,
        "created_at": assignment.created_at,
        "updated_at": assignment.updated_at,
        "words": [
            {
                "id": w.id,
                "word_text": w.word_text,
                "order_index": w.order_index
            }
            for w in sorted(assignment.words, key=lambda x: x.order_index)
        ],
        "word_count": word_count,
        "student_count": student_count
    }


def build_student_assignment_response(assignment: Assignment, student_id: int, db: Session) -> dict:
    """Build assignment response from student's perspective"""
    # Get teacher info
    teacher = db.query(User).filter(User.id == assignment.teacher_id).first()
    teacher_name = teacher.username if teacher else "Unknown"

    # Get word database name
    word_database_name = None
    if assignment.word_database_id:
        database = db.query(WordDatabase).filter(WordDatabase.id == assignment.word_database_id).first()
        if database:
            word_database_name = database.name

    # Get assignment student record
    assignment_student = db.query(AssignmentStudent).filter(
        AssignmentStudent.assignment_id == assignment.id,
        AssignmentStudent.student_id == student_id
    ).first()

    # Calculate progress
    total_words = len(assignment.words)
    completed_words = db.query(AssignmentSubmission).filter(
        AssignmentSubmission.assignment_id == assignment.id,
        AssignmentSubmission.student_id == student_id
    ).count()

    completion_percentage = (completed_words / total_words * 100) if total_words > 0 else 0

    # Check if overdue
    is_overdue = False
    if assignment.due_date and datetime.utcnow() > assignment.due_date.replace(tzinfo=None):
        is_overdue = True

    return {
        "id": assignment.id,
        "title": assignment.title,
        "description": assignment.description,
        "mode": assignment.mode or "practice",
        "teacher_name": teacher_name,
        "word_database_name": word_database_name,
        "due_date": assignment.due_date,
        "assigned_at": assignment_student.assigned_at if assignment_student else None,
        "completed_at": assignment_student.completed_at if assignment_student else None,
        "words": [
            {
                "id": w.id,
                "word_text": w.word_text,
                "order_index": w.order_index
            }
            for w in sorted(assignment.words, key=lambda x: x.order_index)
        ],
        "total_words": total_words,
        "completed_words": completed_words,
        "completion_percentage": round(completion_percentage, 1),
        "is_overdue": is_overdue
    }
